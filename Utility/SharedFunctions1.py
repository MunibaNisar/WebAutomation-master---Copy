import configparser
import string
import time
from datetime import datetime
import random
import os, pypyodbc
import platform
import openpyxl
from Utility.WebDriver import WebDriver
import openpyxl.styles.alignment as XLStyle
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image
from Utility.InputFile import Inputs

input = Inputs()



class commonFunctions:
    # root = input.current_folder
    # get_parent_path = os.path.abspath(os.path.join(root, os.pardir))
    # OutPutFilePath = os.path.join(get_parent_path, 'webopr.xlsx')
    # OutPutFilePath = "E:\Muniba Data\OPR\WebAutomation\webopr.xlsx"
    OutPutFilePath = input.OutPutFilePath
    ExecutionDate = str(datetime.now().date())
    ExecutionTime = str(time.strftime("%H:%M:%S", time.localtime()))
    platformsystem = str(platform.system())
    platformrelease = str(platform.release())
    WindowServer = str(platformsystem + platformrelease)
    SystemUser = os.getlogin()

    # def read_data_from_config_file(self, section, key):
    #     get_parent_path = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
    #     file = os.path.join(get_parent_path, 'config.ini')
    #     config = configparser.ConfigParser()
    #     config.read(file)
    #
    #     value = config[section][key]
    #
    #     return value

    def UpdateExcelTestCase(self, Sheetname, TestCaseID, status,sc_path):  # sc_path

        wb = openpyxl.load_workbook(self.OutPutFilePath)
        wb.sheetnames
        ws = wb[Sheetname]

        first_column = ws['B']
        # del Parameters["AuthToken"]
        for x in range(len(first_column)):
            if (first_column[x].value) == TestCaseID:

                ws.cell(row=x + 1, column=6).alignment = XLStyle.Alignment(horizontal='center', vertical='center',
                                                                           wrap_text=True, wrapText=True)

                ws.cell(row=x + 1, column=7).value = self.ExecutionDate
                # # ws.cell(row=x+1 , column=8).value = self.ExecutionTime
                # ws.cell(row=x + 1, column=8).value = str(time.strftime("%H:%M:%S", time.localtime()))
                # ProcessingTime = float(str((time.process_time() - starttime + 2)))
                # ws.cell(row=x + 1, column=10).value = ProcessingTime
                ws.cell(row=x + 1, column=8).value = self.SystemUser
                ws.cell(row=x + 1, column=9).value = self.WindowServer
                ws.cell(row=x + 1, column=11).value = status

                if status == 'Passed':

                    ws.cell(row=x + 1, column=11).fill = PatternFill(start_color='92D050', end_color='92D050',
                                                                     fill_type='solid')
                else:
                    ws.cell(row=x + 1, column=11).fill = PatternFill(start_color='FF0000', end_color='FF0000',
                                                                     fill_type='solid')
                    # print("image take")
                    # img = openpyxl.drawing.image.Image(sc_path)
                    # img.width = 150
                    # img.height = 100
                    # # define the cell and row position of screen shots in excel
                    # cell = self.cell_position(TeestCaseID)
                    # cell = "N" + str(cell)
                    # # add image in excel file
                    # print("image adding")
                    # ws.add_image(img, cell)
                    # print("image added")
                wb.save('' + self.OutPutFilePath + '')

        print("-------------------Test Results------------------")
        print("Test Case ID : " + TestCaseID)
        # print("URL  : " + URL)
        print("Header  : ")
        # print(Parameters)
        print("Body  : ")
        # print(data)
        print("Test Status  : " + status)
        print("Response  : ")
        # print(resp)
        print("--------------------------------------------")

    def screenshort(self, test_case_id, driver):
        # root = input.current_folder
        # get_parent_path = os.path.abspath(os.path.join(root, os.pardir))
        # OutPuSCPath = os.path.join(get_parent_path, 'Screenshots')
        OutPuSCPath = input.OutPuSCPath
        sc_name = test_case_id + ".png"
        save_path_sc = os.path.join(OutPuSCPath, sc_name)
        driver.save_screenshot(save_path_sc)
        return save_path_sc

    def cell_position(self, testcase_id):
        TC_no = testcase_id.split('-')
        cel = TC_no[-1]
        if cel.startswith('0'):
            cell = int(cel[-1]) + 9
            print(cell)
        else:
            cell = int(cel) + 9
        return cell

    def driver_initialization(self):
        driver_initialization = WebDriver()
        driver = driver_initialization.browser_startup('chrome')
        return driver

    def GenrateSimpleStringLimit10(self):

        Simplestring = "".join([random.choice(string.ascii_uppercase) for x in range(6)])
        SimpleString = '' + Simplestring + 'test'

        return SimpleString

    def GenrateValidIPString(self):

        '''Generate only integers Or IP Address'''
        first = str(random.randrange(100, 255))
        second = str(random.randint(1, 255))
        third = str(random.randint(1, 255))
        fourth = str(random.randint(1, 255))
        ValidIP = str('' + first + '.' + second + '.' + third + '.' + fourth + '')

        return ValidIP

